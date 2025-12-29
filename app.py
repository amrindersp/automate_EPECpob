import os
import uuid
from io import BytesIO

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    send_file,
    session,
)
import pandas as pd
from openpyxl.styles import PatternFill

# ---------------- BASIC SETUP ----------------

app = Flask(__name__)

# Secret key for sessions (use an env var in production)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key")

# Folder to store per-session temporary Excel files
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TMP_DIR = os.path.join(BASE_DIR, "tmp")
os.makedirs(TMP_DIR, exist_ok=True)


# --------- Utility: critical cleaning for NED columns ---------

def clean_ned_column(df, col_name):
    """
    Critical cleaning rules for NED Pass Numbers:
    - Treat strictly as text
    - Remove leading/trailing spaces
    - Remove empty / nan / none / null / na / n/a
    - Do not assume fixed NED pattern
    - Automatically ignore footer-type rows:
      * value has no digits at all
      * row lies in the last 15 rows of the file
    """
    s = df[col_name].astype(str).str.strip()

    empty_markers = {"", "nan", "none", "null", "na", "n/a"}
    mask_empty_like = s.str.lower().isin(empty_markers)

    # Values that have no digits at all (pure text)
    mask_no_digits = ~s.str.contains(r"\d", regex=True)

    # Consider last 15 rows as footer zone
    last_n = 15
    footer_start_index = max(0, len(df) - last_n)
    approx_footer = df.index >= footer_start_index

    # Footer-type rows: pure text in footer area
    mask_footer_text = mask_no_digits & approx_footer

    keep_mask = ~(mask_empty_like | mask_footer_text)

    df_clean = df.loc[keep_mask].copy()
    df_clean[col_name] = s.loc[keep_mask]

    return df_clean


# --------- Helpers to read/write per-session files and data ---------

def get_job_id():
    """Ensure each browser session has a unique job id."""
    if "job_id" not in session:
        session["job_id"] = str(uuid.uuid4())
    return session["job_id"]


def get_path(name):
    """Build a file path in tmp folder for this job."""
    job_id = get_job_id()
    filename = f"{job_id}_{name}.xlsx"
    return os.path.join(TMP_DIR, filename)


def save_df(df, name):
    """Save a dataframe for this session."""
    path = get_path(name)
    df.to_excel(path, index=False)
    return path


def load_df(name):
    """Load a dataframe for this session."""
    path = get_path(name)
    return pd.read_excel(path)


def save_duplicate_workbook(output_bytesio):
    """Save duplicate-highlight workbook for this session."""
    job_id = get_job_id()
    filename = f"{job_id}_duplicates.xlsx"
    path = os.path.join(TMP_DIR, filename)
    with open(path, "wb") as f:
        f.write(output_bytesio.getvalue())
    session["duplicate_path"] = path


# ---------------- STEP 1 : UPLOAD ----------------

@app.route("/", methods=["GET", "POST"])
def upload():
    if request.method == "POST":
        # Reset job and simple session data
        session.clear()
        job_id = get_job_id()

        pob_file = request.files["pob"]
        portal_file = request.files["portal"]

        pob_df = pd.read_excel(pob_file)
        portal_df = pd.read_excel(portal_file)

        # Store column names only (lightweight) in session
        session["pob_cols"] = list(pob_df.columns)
        session["portal_cols"] = list(portal_df.columns)

        # Save raw dataframes per session
        save_df(pob_df, "pob_raw")
        save_df(portal_df, "portal_raw")

        return redirect(url_for("select_columns"))

    return render_template("upload.html")


# ---------------- STEP 2 : COLUMN SELECTION ----------------

@app.route("/select_columns", methods=["GET", "POST"])
def select_columns():
    pob_cols = session.get("pob_cols")
    portal_cols = session.get("portal_cols")

    if pob_cols is None or portal_cols is None:
        return redirect(url_for("upload"))

    if request.method == "POST":
        session["pob_ned"] = request.form["pob_ned"]
        session["pob_name"] = request.form["pob_name"]
        session["portal_ned"] = request.form["portal_ned"]
        session["portal_name"] = request.form["portal_name"]

        # Load raw data, clean selected NED columns, save cleaned versions
        pob_df = load_df("pob_raw")
        portal_df = load_df("portal_raw")

        pob_ned = session["pob_ned"]
        portal_ned = session["portal_ned"]

        pob_df_clean = clean_ned_column(pob_df, pob_ned)
        portal_df_clean = clean_ned_column(portal_df, portal_ned)

        save_df(pob_df_clean, "pob_clean")
        save_df(portal_df_clean, "portal_clean")

        return redirect(url_for("check_duplicates"))

    return render_template(
        "column_select.html",
        pob_cols=pob_cols,
        portal_cols=portal_cols,
    )


# ---------------- STEP 3 : DUPLICATE CHECK ----------------

@app.route("/check_duplicates")
def check_duplicates():
    pob_ned = session.get("pob_ned")
    portal_ned = session.get("portal_ned")

    if pob_ned is None or portal_ned is None:
        return redirect(url_for("upload"))

    pob = load_df("pob_clean")
    portal = load_df("portal_clean")

    pob_dup_mask = pob[pob_ned].duplicated(keep=False)
    portal_dup_mask = portal[portal_ned].duplicated(keep=False)

    pob_dup = pob_dup_mask.any()
    portal_dup = portal_dup_mask.any()

    if pob_dup or portal_dup:
        # Build Excel file with duplicate NED cells highlighted
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            pob.to_excel(writer, sheet_name="POB", index=False)
            portal.to_excel(writer, sheet_name="PORTAL", index=False)

            wb = writer.book
            pob_ws = wb["POB"]
            portal_ws = wb["PORTAL"]

            yellow_fill = PatternFill(
                start_color="FFFF00",
                end_color="FFFF00",
                fill_type="solid",
            )

            # POB duplicates
            pob_ned_idx = pob.columns.get_loc(pob_ned) + 1
            for i, is_dup in enumerate(pob_dup_mask, start=2):
                if is_dup:
                    pob_ws.cell(row=i, column=pob_ned_idx).fill = yellow_fill

            # PORTAL duplicates
            portal_ned_idx = portal.columns.get_loc(portal_ned) + 1
            for i, is_dup in enumerate(portal_dup_mask, start=2):
                if is_dup:
                    portal_ws.cell(row=i, column=portal_ned_idx).fill = yellow_fill

        output.seek(0)
        save_duplicate_workbook(output)

        return render_template(
            "duplicate_warning.html",
            pob_dup=pob_dup,
            portal_dup=portal_dup,
        )

    return redirect(url_for("user_inputs"))


@app.route("/duplicate_decision", methods=["POST"])
def duplicate_decision():
    if request.form["decision"] == "reupload":
        return redirect(url_for("upload"))
    return redirect(url_for("user_inputs"))


@app.route("/download_duplicates")
def download_duplicates():
    path = session.get("duplicate_path")
    if not path or not os.path.exists(path):
        return redirect(url_for("upload"))
    return send_file(
        path,
        download_name="Uploaded_with_Duplicates_Highlighted.xlsx",
        as_attachment=True,
    )


# ---------------- STEP 4 : USER INPUTS ----------------

@app.route("/user_inputs", methods=["GET", "POST"])
def user_inputs():
    # Ensure data uploaded and columns selected
    if "pob_ned" not in session or "portal_ned" not in session:
        return redirect(url_for("upload"))

    if request.method == "POST":
        session["inputs"] = request.form.to_dict()
        return redirect(url_for("generate"))
    return render_template("user_inputs.html")


# ---------------- STEP 5 : PROCESS DATA ----------------

@app.route("/generate")
def generate():
    if "inputs" not in session:
        return redirect(url_for("upload"))

    inputs = session["inputs"]
    pob_ned = session["pob_ned"]
    portal_ned = session["portal_ned"]
    pob_name = session["pob_name"]
    portal_name = session["portal_name"]

    pob = load_df("pob_clean")
    portal = load_df("portal_clean")

    missing_in_portal = pob[~pob[pob_ned].isin(portal[portal_ned])]
    missing_in_pob = portal[~portal[portal_ned].isin(pob[pob_ned])]

    manifest_count = len(missing_in_portal)
    return_count = len(missing_in_pob)

    session["manifest_count"] = manifest_count
    session["return_count"] = return_count

    # RFM
    rfm = pd.DataFrame({
        "Passenger Category": inputs["rfm_category"],
        "NED Pass No.": missing_in_portal[pob_ned],
        "Travelling Vendor Code": inputs["vendor_code"],
        "Vendor Name": inputs["vendor_name"],
        "Vendor Employee Name": missing_in_portal[pob_name],
        "Gender": inputs["gender"],
        "Designation": "",
        "Originating Point": inputs["rfm_origin"],
        "Destination Point": inputs["rfm_destination"],
        "": "",
        "Charge": inputs["charge"],
    })

    # Manifest
    manifest = pd.DataFrame({
        "Passenger Weight": inputs["passenger_weight"],
        "Baggage Weight": inputs["baggage_weight"],
        "": "",
        " ": "",
        "  ": "",
        "Time Reported": inputs["time_reported"],
    }, index=rfm.index)

    # Return Manifest
    return_manifest = pd.DataFrame({
        "Passenger Category": inputs["return_category"],
        "Smart Card No.": missing_in_pob[portal_ned],
        "Supplier": inputs["supplier"],
        "Vendor Employee Name": missing_in_pob[portal_name],
        "Gender": inputs["gender"],
        "Designation": "",
        "": "",
        "Charge": inputs["charge"],
        "Pax wt.": inputs["passenger_weight"],
        "Baggage": inputs["baggage_weight"],
        "Originating Point": inputs["return_origin"],
        "Destination Point": inputs["return_destination"],
    })

    # Save final dataframes for this session
    save_df(rfm, "rfm_final")
    save_df(manifest, "manifest_final")
    save_df(return_manifest, "return_manifest_final")

    return render_template(
        "result.html",
        manifest_count=manifest_count,
        return_count=return_count,
    )


# ---------------- STEP 6 : DOWNLOAD ----------------

@app.route("/download")
def download():
    # Load per-session final dataframes
    rfm = load_df("rfm_final")
    manifest = load_df("manifest_final")
    return_manifest = load_df("return_manifest_final")

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        rfm.to_excel(writer, index=False, sheet_name="RFM")
        manifest.to_excel(writer, index=False, sheet_name="Manifest")
        return_manifest.to_excel(writer, index=False, sheet_name="Return Manifest")
    output.seek(0)

    return send_file(
        output,
        download_name="Final_Output.xlsx",
        as_attachment=True,
    )


if __name__ == "__main__":
    app.run(debug=True)
